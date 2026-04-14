#!/usr/bin/env python3
"""Build a vocabulary deck from local PDF/XLS resources.

Default source directory is aligned with the user's OneDrive path.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

import openpyxl
from deep_translator import GoogleTranslator
from pypdf import PdfReader

DEFAULT_SOURCE_DIR = Path(
    "/Users/yh/Library/CloudStorage/OneDrive-Personal/01-Kardiology E kitap/ingilizce"
)
DEFAULT_OUTPUT = Path(__file__).resolve().parents[1] / "public" / "data" / "cards.generated.json"

PHRASAL_PARTICLES = {
    "up",
    "down",
    "out",
    "in",
    "on",
    "off",
    "over",
    "into",
    "away",
    "back",
    "around",
    "through",
    "with",
    "for",
    "from",
    "by",
    "at",
}

IGNORED_TERMS = {
    "authors",
    "author",
    "friends and family",
    "appearance and background",
    "personality traits",
}

ENTRY_RE = re.compile(r"^([A-Za-z][A-Za-z'\- ]{1,60})\s*\(([^)]+)\)\s*:\s*(.+)$")
ESSENTIAL_HEAD_RE = re.compile(r"^([A-Za-z][A-Za-z'\- ]{1,40})\s*\[[^\]]+\]\s*([a-z]{1,5})\.?$", re.I)
BULLET_RE = re.compile(r"^[-–—*•»]+\s*(.+)$")
NON_ASCII_RE = re.compile(r"[^\x00-\x7F]")
SPACE_RE = re.compile(r"\s+")
YDS_ENTRY_RE = re.compile(r"^\d+\)\s*([A-Za-z][A-Za-z /'\-]{1,60}?)\s*-\s*(.+)$")


@dataclass
class Card:
    term: str
    type: str
    register: str
    meaning_en: str
    meaning_tr: str
    examples: list[str]
    source: str
    part_of_speech: str = ""
    tags: list[str] = field(default_factory=list)
    normalized: str = ""
    card_id: str = ""

    def finalize(self) -> None:
        self.term = clean_text(self.term)
        self.meaning_en = clean_text(self.meaning_en)
        self.meaning_tr = clean_text(self.meaning_tr)
        self.examples = [clean_text(example) for example in self.examples if clean_text(example)]
        self.examples = dedupe_list(self.examples)[:4]
        self.tags = dedupe_list([clean_text(tag) for tag in self.tags if clean_text(tag)])
        self.normalized = normalize_text(self.term)
        hash_src = f"{self.normalized}|{self.type}|{self.register}"
        digest = hashlib.md5(hash_src.encode("utf-8")).hexdigest()[:10]
        slug = self.normalized.replace(" ", "-")[:40].strip("-") or "card"
        self.card_id = f"{slug}-{digest}"

    def to_dict(self) -> dict:
        return {
            "id": self.card_id,
            "term": self.term,
            "normalized": self.normalized,
            "type": self.type,
            "register": self.register,
            "partOfSpeech": self.part_of_speech,
            "meaningEn": self.meaning_en,
            "meaningTr": self.meaning_tr,
            "examples": self.examples,
            "source": self.source,
            "tags": self.tags,
        }


def clean_text(text: str) -> str:
    text = text.replace("\u00ad", "")
    text = text.replace("’", "'").replace("`", "'")
    text = text.replace("“", '"').replace("”", '"')
    text = text.replace("ı", "i")
    text = SPACE_RE.sub(" ", text)
    return text.strip(" -\t\n\r")


def normalize_text(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    ascii_text = normalized.encode("ascii", "ignore").decode("ascii")
    ascii_text = re.sub(r"[^a-zA-Z0-9 ]+", " ", ascii_text.lower())
    return SPACE_RE.sub(" ", ascii_text).strip()


def dedupe_list(values: Iterable[str]) -> list[str]:
    seen = set()
    out: list[str] = []
    for value in values:
        key = normalize_text(value)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(value)
    return out


def contains_turkish(text: str) -> bool:
    lower = text.lower()
    if any(ch in lower for ch in "ığüşöç"):
        return True
    return bool(re.search(r"\b(mek|mak|lik|lık|dir|dır|tir|tır|iyor|yor)\b", lower))


def detect_type(term: str, part_of_speech: str) -> str:
    normalized = normalize_text(term)
    tokens = normalized.split()
    pos = part_of_speech.lower()

    if "idiom" in pos:
        return "idiom"

    if "phr" in pos or "phrasal" in pos:
        return "phrasal_verb"

    if len(tokens) >= 2:
        if any(token in PHRASAL_PARTICLES for token in tokens[1:]):
            return "phrasal_verb"
        if len(tokens) >= 3:
            return "idiom"
        return "collocation"

    return "word"


def split_en_tr(definition: str) -> tuple[str, str]:
    definition = clean_text(definition)
    match = re.search(r"\(([^()]{2,240})\)\s*$", definition)
    if not match:
        return definition, ""

    candidate = clean_text(match.group(1))
    if not contains_turkish(candidate):
        return definition, ""

    en = clean_text(definition[: match.start()])
    return en, candidate


def parse_vocab_phrasal_pdf(path: Path, academic_terms: set[str]) -> list[Card]:
    cards: list[Card] = []
    reader = PdfReader(str(path))

    current: dict | None = None

    def flush_current() -> None:
        nonlocal current
        if not current:
            return

        term = clean_text(current["term"])
        if normalize_text(term) in IGNORED_TERMS:
            current = None
            return

        raw_definition = clean_text(" ".join(current["def_lines"]))
        raw_definition = re.sub(r"\bsyn:\b.*", "", raw_definition, flags=re.I)
        meaning_en, meaning_tr = split_en_tr(raw_definition)

        if len(term) < 2 or len(meaning_en) < 5:
            current = None
            return

        register = "academic"
        tokens = set(normalize_text(term).split())
        if tokens and tokens.isdisjoint(academic_terms):
            register = "daily"

        part_of_speech = clean_text(current["pos"])[:30]
        card = Card(
            term=term,
            type=detect_type(term, part_of_speech),
            register=register,
            meaning_en=meaning_en,
            meaning_tr=meaning_tr,
            examples=current["examples"],
            source=path.name,
            part_of_speech=part_of_speech,
            tags=["source:pdf", "book:vocabulary+phrasal"],
        )
        card.finalize()

        if card.normalized and card.meaning_en:
            cards.append(card)

        current = None

    for page in reader.pages:
        text = page.extract_text() or ""
        lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]

        for line in lines:
            if ".indd" in line or line.isdigit():
                continue

            entry_match = ENTRY_RE.match(line)
            if entry_match:
                flush_current()
                current = {
                    "term": entry_match.group(1),
                    "pos": entry_match.group(2),
                    "def_lines": [entry_match.group(3)],
                    "examples": [],
                }
                continue

            if not current:
                continue

            bullet_match = BULLET_RE.match(line)
            if bullet_match:
                sentence = clean_text(bullet_match.group(1))
                if len(sentence.split()) >= 4:
                    current["examples"].append(sentence)
                continue

            if re.match(r"^[A-Z][A-Z ,&/\-]{4,}$", line):
                continue

            if line.lower().startswith(("syn:", "see also:")):
                continue

            if len(line.split()) >= 3:
                current["def_lines"].append(line)

    flush_current()
    return cards


def parse_essential_pdf(path: Path) -> list[Card]:
    cards: list[Card] = []
    reader = PdfReader(str(path))

    current: dict | None = None

    def flush_current() -> None:
        nonlocal current
        if not current:
            return

        term = clean_text(current["term"])
        definition = clean_text(current["definition"])

        if len(term) < 2 or len(definition) < 8:
            current = None
            return

        card = Card(
            term=term,
            type=detect_type(term, current["pos"]),
            register="daily",
            meaning_en=definition,
            meaning_tr="",
            examples=current["examples"],
            source=path.name,
            part_of_speech=current["pos"],
            tags=["source:pdf", "book:4000-essential"],
        )
        card.finalize()

        if card.normalized and card.meaning_en:
            cards.append(card)

        current = None

    for page in reader.pages:
        text = page.extract_text() or ""
        lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]

        for line in lines:
            sanitized = re.sub(r"^[^A-Za-z]+", "", line)
            head_match = ESSENTIAL_HEAD_RE.match(sanitized)
            if head_match:
                flush_current()
                current = {
                    "term": head_match.group(1),
                    "pos": head_match.group(2),
                    "definition": "",
                    "examples": [],
                }
                continue

            if not current:
                continue

            if line.lower().startswith(("check", "choose", "answer", "fill in", "listen")):
                continue

            bullet_match = BULLET_RE.match(line)
            if bullet_match:
                sentence = clean_text(bullet_match.group(1))
                if len(sentence.split()) >= 4:
                    current["examples"].append(sentence)
                continue

            lower_line = line.lower()
            if not current["definition"] and (
                lower_line.startswith(("to ", "a ", "an ")) or " means " in lower_line
            ):
                current["definition"] = line

    flush_current()
    return cards


def parse_yds_pdf(path: Path) -> list[Card]:
    """Parse YDS style lists: '1) TERM - Example sentence (TÜRKÇE ANLAM)'."""
    cards: list[Card] = []
    reader = PdfReader(str(path))

    # Collect all text, then join multiline entries
    all_text = ""
    for page in reader.pages:
        try:
            all_text += (page.extract_text() or "") + "\n"
        except Exception:
            continue

    # Join continuation lines (lines not starting with a number)
    joined_lines: list[str] = []
    for line in all_text.splitlines():
        line = clean_text(line)
        if not line:
            continue
        if re.match(r"^\d+\)", line):
            joined_lines.append(line)
        elif joined_lines:
            joined_lines[-1] += " " + line

    for line in joined_lines:
        match = YDS_ENTRY_RE.match(line)
        if not match:
            continue

        term = clean_text(match.group(1))
        rest = clean_text(match.group(2))

        # Extract Turkish meaning from last parenthesized block
        meaning_tr = ""
        tr_match = re.search(r"\(([^()]{2,80})\)\s*$", rest)
        if not tr_match:
            # Try finding uppercase Turkish in parens anywhere
            tr_match = re.search(r"\(([A-ZÇĞIİÖŞÜ][A-ZÇĞIİÖŞÜa-zçğıiöşü /-]{1,80})\)", rest)
        if tr_match:
            meaning_tr = clean_text(tr_match.group(1))
            rest = clean_text(rest[:tr_match.start()])

        # Rest is the example sentence
        example = rest.strip(" .")
        if len(term) < 2:
            continue

        card = Card(
            term=term,
            type=detect_type(term, ""),
            register="academic",
            meaning_en="",
            meaning_tr=meaning_tr,
            examples=[example] if len(example) > 10 else [],
            source=path.name,
            tags=["source:pdf", "list:yds-target"],
        )
        card.finalize()
        cards.append(card)
    return cards






def parse_cesur_ozturk_pdf(path: Path) -> list[Card]:
    """Parse Cesur Ozturk Academic Vocabulary. Simple line-based parser."""
    cards: list[Card] = []
    reader = PdfReader(str(path))
    tr_chars = set("ğüşöçıİĞÜŞÖÇ")
    term_re = re.compile(r"^([a-z][a-z ()/'\-]{1,35})\s+([a-z].*)")
    skip_words = {"exercise", "basic", "essential", "verbs", "nouns", "adjective",
                  "sozcuk", "sozciik", "kar", "complete", "match", "fill", "choose"}

    for page in reader.pages:
        try:
            text = page.extract_text() or ""
        except Exception:
            continue
        lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
        i = 0
        while i < len(lines):
            line = lines[i]
            first_word = line.split()[0].lower() if line.split() else ""
            if first_word in skip_words or line.startswith("___") or re.match(r"^\d+\.", line):
                i += 1
                continue
            match = term_re.match(line)
            if not match:
                i += 1
                continue
            term = clean_text(match.group(1))
            rest = clean_text(match.group(2))
            meaning_en = rest
            meaning_tr = ""
            examples: list[str] = []
            i += 1
            for _ in range(6):
                if i >= len(lines):
                    break
                nxt = lines[i]
                if term_re.match(nxt) and not any(c in nxt for c in tr_chars):
                    break
                has_tr = any(c in nxt for c in tr_chars)
                is_sentence = len(nxt.split()) >= 5
                if has_tr and not is_sentence:
                    meaning_tr = nxt
                elif is_sentence and nxt[0].isupper():
                    examples.append(nxt)
                i += 1
            if len(term) < 2 or normalize_text(term) in IGNORED_TERMS:
                continue
            card = Card(
                term=term,
                type=detect_type(term, ""),
                register="academic",
                meaning_en=meaning_en[:200],
                meaning_tr=meaning_tr[:200],
                examples=examples[:2],
                source=path.name,
                tags=["source:pdf", "book:cesur-ozturk"],
            )
            card.finalize()
            if card.normalized and (card.meaning_en or card.meaning_tr):
                cards.append(card)
    return cards

def parse_acl_xls(path: Path) -> list[Card]:
    cards: list[Card] = []

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir) / "acl.xlsx"
        shutil.copy2(path, tmp_path)
        wb = openpyxl.load_workbook(tmp_path, data_only=True)

    ws = wb[wb.sheetnames[0]]

    for row_idx in range(3, ws.max_row + 1):
        addition_1 = ws.cell(row_idx, 2).value
        comp_1 = ws.cell(row_idx, 3).value
        pos_1 = ws.cell(row_idx, 4).value
        comp_2 = ws.cell(row_idx, 5).value
        pos_2 = ws.cell(row_idx, 6).value
        addition_2 = ws.cell(row_idx, 7).value

        parts = [addition_1, comp_1, comp_2, addition_2]
        term = clean_text(" ".join(str(part) for part in parts if part))
        term = re.sub(r"\(\s+", "(", term)
        term = re.sub(r"\s+\)", ")", term)

        if not term or len(term) < 3:
            continue

        part_of_speech = clean_text("/".join(str(part) for part in [pos_1, pos_2] if part))
        meaning_en = "Academic collocation frequently used in formal writing and exam passages."
        example = f"In an academic report, writers often use \"{term}\" for precision."

        card = Card(
            term=term,
            type="collocation",
            register="academic",
            meaning_en=meaning_en,
            meaning_tr="",
            examples=[example],
            source=path.name,
            part_of_speech=part_of_speech,
            tags=["source:xls", "list:academic-collocations"],
        )
        card.finalize()

        if card.normalized:
            cards.append(card)

    return cards


def idiom_seed_cards() -> list[Card]:
    raw = [
        (
            "break the ice",
            "To start a conversation and make people feel less uncomfortable.",
            "ortami yumusatmak, ilk girisimi yapmak",
            "The teacher told a joke to break the ice on the first day.",
        ),
        (
            "hit the books",
            "To study hard, especially for an exam.",
            "derse gomulmek, sikı calismak",
            "I need to hit the books before tomorrow's quiz.",
        ),
        (
            "under the weather",
            "Feeling slightly ill.",
            "keyifsiz veya hasta hissetmek",
            "I felt under the weather, so I stayed home.",
        ),
        (
            "once in a blue moon",
            "Very rarely.",
            "cok nadiren",
            "We eat out once in a blue moon these days.",
        ),
        (
            "on the same page",
            "To share the same understanding or opinion.",
            "ayni fikirde olmak",
            "Before the meeting, let's make sure we are on the same page.",
        ),
        (
            "a piece of cake",
            "Something very easy.",
            "cok kolay bir sey",
            "After enough practice, the final test was a piece of cake.",
        ),
        (
            "burn the midnight oil",
            "To work late into the night.",
            "gece gec saate kadar calismak",
            "She burned the midnight oil to finish her paper.",
        ),
        (
            "cost an arm and a leg",
            "To be very expensive.",
            "cok pahali olmak",
            "New medical devices can cost an arm and a leg.",
        ),
        (
            "get the hang of",
            "To learn how to do something properly.",
            "bir seyin mantigini kavramak",
            "It took me a week to get the hang of pronunciation drills.",
        ),
        (
            "keep an eye on",
            "To watch something carefully.",
            "goz kulak olmak",
            "Can you keep an eye on my notes while I take a break?",
        ),
    ]

    cards: list[Card] = []
    for term, meaning_en, meaning_tr, example in raw:
        card = Card(
            term=term,
            type="idiom",
            register="daily",
            meaning_en=meaning_en,
            meaning_tr=meaning_tr,
            examples=[example],
            source="seed-idioms",
            part_of_speech="idiom",
            tags=["seed"],
        )
        card.finalize()
        cards.append(card)

    return cards


def merge_cards(cards: Iterable[Card]) -> list[Card]:
    merged: dict[tuple[str, str, str], Card] = {}

    def score(card: Card) -> float:
        return (
            len(card.meaning_en) * 0.2
            + len(card.meaning_tr) * 0.2
            + len(card.examples) * 8
            + (5 if card.source.endswith(".xls") else 0)
        )

    for card in cards:
        key = (card.normalized, card.type, card.register)
        existing = merged.get(key)
        if not existing:
            merged[key] = card
            continue

        better = card if score(card) > score(existing) else existing
        weaker = existing if better is card else card

        if not better.meaning_tr and weaker.meaning_tr:
            better.meaning_tr = weaker.meaning_tr

        better.examples = dedupe_list([*better.examples, *weaker.examples])[:4]
        better.tags = dedupe_list([*better.tags, *weaker.tags])
        better.finalize()
        merged[key] = better

    return list(merged.values())


def fill_missing_tr(cards: list[Card], max_translate: int) -> None:
    translator = GoogleTranslator(source="en", target="tr")
    cache: dict[str, str] = {}
    translated = 0

    for card in cards:
        if card.meaning_tr:
            continue
        if translated >= max_translate:
            break

        text_for_translation = card.term if len(card.term) < 70 else card.meaning_en
        key = normalize_text(text_for_translation)
        if not key:
            continue

        if key not in cache:
            try:
                cache[key] = clean_text(translator.translate(text_for_translation))
            except Exception:
                cache[key] = ""

        card.meaning_tr = cache[key]
        card.finalize()

        if card.meaning_tr:
            translated += 1

    for card in cards:
        if card.meaning_tr:
            continue
        card.meaning_tr = card.term
        card.finalize()


def apply_quota(cards: list[Card], max_daily: int, max_academic: int) -> list[Card]:
    daily = [card for card in cards if card.register == "daily"]
    academic = [card for card in cards if card.register == "academic"]

    def sort_cards(items: list[Card]) -> list[Card]:
        return sorted(items, key=lambda card: (card.term.lower(), card.source.lower()))

    daily_sorted = sort_cards(daily)
    academic_sorted = sort_cards(academic)

    daily_phrasal = sort_cards([card for card in daily_sorted if card.type == "phrasal_verb"])
    daily_idioms = sort_cards([card for card in daily_sorted if card.type == "idiom"])
    daily_other = sort_cards(
        [card for card in daily_sorted if card.type not in {"phrasal_verb", "idiom"}]
    )

    phrasal_target = min(len(daily_phrasal), max(20, max_daily // 4))
    idiom_target = min(len(daily_idioms), max(15, max_daily // 6))

    selected_daily = daily_phrasal[:phrasal_target] + daily_idioms[:idiom_target]

    for card in daily_other:
        if len(selected_daily) >= max_daily:
            break
        selected_daily.append(card)

    if len(selected_daily) < max_daily:
        extras = [
            card
            for card in daily_sorted
            if card.card_id not in {picked.card_id for picked in selected_daily}
        ]
        selected_daily.extend(extras[: max_daily - len(selected_daily)])

    selected_academic = academic_sorted[:max_academic]
    selected = selected_daily[:max_daily] + selected_academic
    return sorted(selected, key=lambda card: (card.register, card.term.lower()))



def build_deck(
    source_dir: Path,
    output_path: Path,
    max_daily: int,
    max_academic: int,
    max_translate: int,
) -> None:
    if not source_dir.exists():
        raise FileNotFoundError(f"Source directory does not exist: {source_dir}")

    acl_file = source_dir / "2021_Teachers_AcademicCollocationList.xls"
    vocab_phrasal_file = source_dir / "English Vocabulary and Phrasel Verbs_opt_Optimized.pdf"
    essential_files = sorted(source_dir.glob("Paul Nation_4000 Essential English Words *_2009.pdf"))
    yds_files = sorted(source_dir.glob("*[Yy][Dd][Ss]*")) + sorted(source_dir.glob("*[Yy][Dd][Tt]*")) + sorted(source_dir.glob("*[Kk][Pp][Dd][Ss]*"))
    yds_files = sorted(list({f for f in yds_files if f.suffix.lower() == ".pdf"}))
    cesur_ozturk_file = source_dir / "Essential Academic Vocabulary-Cesur ÖZTÜRK.pdf"

    all_cards: list[Card] = []

    print(f"[1/6] Parsing academic collocations: {acl_file.name}")
    if acl_file.exists():
        acl_cards = parse_acl_xls(acl_file)
        all_cards.extend(acl_cards)
        academic_terms: set[str] = set()
        for card in acl_cards:
            academic_terms.update(card.normalized.split())
        print(f"  -> extracted {len(acl_cards)} academic collocation cards")
    else:
        academic_terms = set()
        print("  [!] File not found, skipping")

    print(f"[2/6] Parsing vocabulary + phrasal verbs")
    if vocab_phrasal_file.exists():
        vocab_cards = parse_vocab_phrasal_pdf(vocab_phrasal_file, academic_terms)
        all_cards.extend(vocab_cards)
        print(f"  -> extracted {len(vocab_cards)} entries")
    else:
        print("  [!] File not found, skipping")

    if essential_files:
        print(f"[3/6] Parsing 4000 essential series ({len(essential_files)} files)")
        for file in essential_files:
            extracted = parse_essential_pdf(file)
            all_cards.extend(extracted)
            print(f"  -> {file.name}: {len(extracted)} entries")
    else:
        print("[3/6] 4000 essential files not found, skipping")

    if yds_files:
        print(f"[4/6] Parsing YDS/YDT/KPDS target lists ({len(yds_files)} files)")
        for file in yds_files:
            extracted = parse_yds_pdf(file)
            all_cards.extend(extracted)
            print(f"  -> {file.name}: {len(extracted)} entries")
    else:
        print("[4/6] No YDS/YDT/KPDS files found, skipping")

    if cesur_ozturk_file.exists():
        print(f"[5/6] Parsing Cesur Öztürk: {cesur_ozturk_file.name}")
        extracted = parse_cesur_ozturk_pdf(cesur_ozturk_file)
        all_cards.extend(extracted)
        print(f"  -> extracted {len(extracted)} entries")
    else:
        print("[5/6] Cesur Öztürk file not found, skipping")

    print("[6/6] Adding idiom starter pack")
    seed_idioms = idiom_seed_cards()
    all_cards.extend(seed_idioms)

    merged = merge_cards(all_cards)

    print(f"  -> merged deck size before quota: {len(merged)}")

    fill_missing_tr(merged, max_translate=max_translate)
    selected = apply_quota(merged, max_daily=max_daily, max_academic=max_academic)

    payload = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceDir": str(source_dir),
        "totalCards": len(selected),
        "stats": {
            "daily": len([card for card in selected if card.register == "daily"]),
            "academic": len([card for card in selected if card.register == "academic"]),
            "phrasal_verbs": len([card for card in selected if card.type == "phrasal_verb"]),
            "idioms": len([card for card in selected if card.type == "idiom"]),
        },
        "cards": [card.to_dict() for card in selected],
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Deck generated -> {output_path}")
    print(
        f"  total={payload['totalCards']} daily={payload['stats']['daily']} "
        f"academic={payload['stats']['academic']} "
        f"phrasal={payload['stats']['phrasal_verbs']} idioms={payload['stats']['idioms']}"
    )

def main() -> None:
    parser = argparse.ArgumentParser(description="Build vocabulary deck from local resources")
    parser.add_argument("--source-dir", type=Path, default=DEFAULT_SOURCE_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--max-daily", type=int, default=1500)
    parser.add_argument("--max-academic", type=int, default=1500)
    parser.add_argument(
        "--max-translate",
        type=int,
        default=600,
        help="Max number of cards to translate with online service before fallback",
    )

    args = parser.parse_args()

    build_deck(
        source_dir=args.source_dir,
        output_path=args.output,
        max_daily=max(100, args.max_daily),
        max_academic=max(100, args.max_academic),
        max_translate=max(0, args.max_translate),
    )


if __name__ == "__main__":
    main()
